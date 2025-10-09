"""用户服务：封装用户信息聚合与用户管理业务逻辑。"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Iterable, Optional

from fastapi import UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.packages.system.core.constants import (
    DEFAULT_ADMIN_USERNAME,
    HTTP_STATUS_BAD_REQUEST,
    HTTP_STATUS_CONFLICT,
    HTTP_STATUS_NOT_FOUND,
    HTTP_STATUS_OK,
)
from app.packages.system.core.enums import UserStatusEnum
from app.packages.system.core.exceptions import AppException
from app.packages.system.core.responses import create_response
from app.packages.system.core.timezone import format_datetime
from app.packages.system.core.security import get_password_hash
from app.packages.system.crud.organizations import organization_crud
from app.packages.system.crud.roles import role_crud
from app.packages.system.crud.users import user_crud
from app.packages.system.models.organization import Organization
from app.packages.system.models.role import Role
from app.packages.system.models.user import User


class UserService:
    """聚合用户相关的核心业务能力。"""

    _STATUS_LABELS = {
        UserStatusEnum.NORMAL.value: "正常",
        UserStatusEnum.DISABLED.value: "停用",
    }

    _IMPORT_HEADERS = ("用户名", "密码", "用户昵称", "状态", "角色", "备注")

    # ------------------------------------------------------------------
    # 个人信息
    # ------------------------------------------------------------------

    def build_user_profile(self, user: User) -> dict:
        """整理用户所属组织、角色与权限，构造统一响应。"""
        organization = None
        if user.organization is not None:
            organization = {
                "org_id": user.organization.id,
                "org_name": user.organization.name,
            }

        roles = [role.name for role in user.roles]
        permission_codes: set[str] = set()
        for role in user.roles:
            permission_codes.update(
                perm.name for perm in role.permissions if getattr(perm, "name", None)
            )
            permission_codes.update(
                item.permission_code
                for item in getattr(role, "access_controls", [])
                if getattr(item, "permission_code", None)
            )
        permissions = sorted(code for code in permission_codes if code)

        data = {
            "user_id": user.id,
            "username": user.username,
            "nickname": user.nickname,
            "status": user.status,
            "organization": organization,
            "roles": roles,
            "permissions": permissions,
        }
        return create_response("获取用户信息成功", data, HTTP_STATUS_OK)

    # ------------------------------------------------------------------
    # 用户管理：查询
    # ------------------------------------------------------------------

    def list_users(
        self,
        db: Session,
        *,
        username: Optional[str] = None,
        statuses: Optional[Iterable[str]] = None,
        start_time: Optional[datetime] = None,
        end_time: Optional[datetime] = None,
        page: int = 1,
        page_size: int = 20,
    ) -> dict:
        page = max(page, 1)
        page_size = max(page_size, 1)
        normalized_statuses = self._normalize_statuses(statuses)

        items, total = user_crud.list_with_filters(
            db,
            username=username,
            statuses=normalized_statuses,
            start_time=start_time,
            end_time=end_time,
            skip=(page - 1) * page_size,
            limit=page_size,
        )

        payload = {
            "total": total,
            "items": [self._serialize_user_summary(item) for item in items],
            "page": page,
            "page_size": page_size,
        }
        return create_response("获取用户列表成功", payload, HTTP_STATUS_OK)

    # ------------------------------------------------------------------
    # 用户管理：增删改
    # ------------------------------------------------------------------

    def create_user(
        self,
        db: Session,
        *,
        username: str,
        password: str,
        nickname: Optional[str] = None,
        status: Optional[str] = None,
        role_ids: Optional[Iterable[int]] = None,
        remark: Optional[str] = None,
        organization_id: Optional[int] = None,
    ) -> dict:
        trimmed_username = (username or "").strip()
        if not trimmed_username:
            raise AppException("用户名不能为空", HTTP_STATUS_BAD_REQUEST)

        if user_crud.get_by_username(db, trimmed_username):
            raise AppException("用户名已存在", HTTP_STATUS_CONFLICT)

        normalized_status = self._normalize_status(status or UserStatusEnum.NORMAL.value)
        roles = self._load_roles(db, role_ids) or []
        organization = self._load_organization(db, organization_id)
        normalized_nickname = self._normalize_optional_text(nickname)
        normalized_remark = self._normalize_optional_text(remark)

        hashed_password = get_password_hash(password)
        user = user_crud.create_with_roles(
            db,
            username=trimmed_username,
            hashed_password=hashed_password,
            nickname=normalized_nickname,
            organization_id=organization.id if organization else None,
            status=normalized_status,
            remark=normalized_remark,
            roles=roles,
        )

        data = self._serialize_user_detail(user)
        return create_response("创建用户成功", data, HTTP_STATUS_OK)

    def update_user(
        self,
        db: Session,
        *,
        user_id: int,
        nickname: Optional[str] = None,
        status: Optional[str] = None,
        role_ids: Optional[Iterable[int]] = None,
        remark: Optional[str] = None,
        organization_id: Optional[int] = None,
    ) -> dict:
        user = user_crud.get(db, user_id)
        if user is None:
            raise AppException("用户不存在或已删除", HTTP_STATUS_NOT_FOUND)

        normalized_status = self._normalize_status(status or user.status)
        roles = self._load_roles(db, role_ids)
        organization = self._load_organization(db, organization_id)
        normalized_nickname = self._normalize_optional_text(nickname)
        normalized_remark = self._normalize_optional_text(remark)

        user.nickname = normalized_nickname
        user.status = normalized_status
        user.is_active = normalized_status == UserStatusEnum.NORMAL.value
        user.remark = normalized_remark
        user.organization_id = organization.id if organization else None
        if roles is not None:
            user.roles = roles

        db.add(user)
        db.commit()
        db.refresh(user)

        data = self._serialize_user_detail(user)
        return create_response("更新用户成功", data, HTTP_STATUS_OK)

    def delete_user(self, db: Session, *, user_id: int) -> dict:
        user = user_crud.get(db, user_id)
        if user is None:
            raise AppException("用户不存在或已删除", HTTP_STATUS_NOT_FOUND)
        if user.username == DEFAULT_ADMIN_USERNAME:
            raise AppException("系统内置管理员不允许删除", HTTP_STATUS_BAD_REQUEST)

        user_crud.soft_delete(db, user)
        payload = {"user_id": user_id}
        return create_response("删除用户成功", payload, HTTP_STATUS_OK)

    def reset_password(self, db: Session, *, user_id: int, new_password: str) -> dict:
        user = user_crud.get(db, user_id)
        if user is None:
            raise AppException("用户不存在或已删除", HTTP_STATUS_NOT_FOUND)

        trimmed = (new_password or "").strip()
        if len(trimmed) < 6:
            raise AppException("密码长度不能少于 6 位", HTTP_STATUS_BAD_REQUEST)

        user.hashed_password = get_password_hash(trimmed)
        db.add(user)
        db.commit()
        db.refresh(user)

        payload = {"user_id": user.id}
        return create_response("重置密码成功", payload, HTTP_STATUS_OK)

    # ------------------------------------------------------------------
    # 用户管理：模板、导入导出
    # ------------------------------------------------------------------

    def export_users(
        self,
        db: Session,
        *,
        username: Optional[str] = None,
        statuses: Optional[Iterable[str]] = None,
        start_time: Optional[datetime] = None,
        end_time: Optional[datetime] = None,
    ) -> StreamingResponse:
        normalized_statuses = self._normalize_statuses(statuses)
        items, _ = user_crud.list_with_filters(
            db,
            username=username,
            statuses=normalized_statuses,
            start_time=start_time,
            end_time=end_time,
            skip=0,
            limit=10_000,
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "用户列表"
        sheet.append(["用户名称", "用户昵称", "状态", "角色", "创建时间", "备注"])

        for item in items:
            sheet.append(
                [
                    item.username,
                    item.nickname or "",
                    self._STATUS_LABELS.get(item.status, item.status),
                    ", ".join(sorted(role.name for role in item.roles)),
                    format_datetime(item.create_time) or "",
                    item.remark or "",
                ]
            )

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
        filename = f"users-{timestamp}.xlsx"
        response = StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        return response

    def download_template(self) -> StreamingResponse:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "用户导入模板"
        sheet.append(list(self._IMPORT_HEADERS))
        sheet.append(["demo", "password123", "演示用户", UserStatusEnum.NORMAL.value, "1", "示例备注，可留空"])

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response.headers["Content-Disposition"] = "attachment; filename=user-template.xlsx"
        return response

    def import_users(self, db: Session, *, file: UploadFile) -> dict:
        content = file.file.read()
        if not content:
            raise AppException("导入文件不能为空", HTTP_STATUS_BAD_REQUEST)

        workbook = load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.active
        header_iter = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
        try:
            first_header = next(header_iter)
        except StopIteration as exc:  # pragma: no cover - 防御性判断
            raise AppException("导入模版不匹配，请下载最新模版", HTTP_STATUS_BAD_REQUEST) from exc
        header_row = [self._normalize_optional_text(cell) for cell in first_header]
        if tuple(header_row[: len(self._IMPORT_HEADERS)]) != self._IMPORT_HEADERS:
            raise AppException("导入模版不匹配，请下载最新模版", HTTP_STATUS_BAD_REQUEST)

        role_lookup = self._prepare_role_lookup(db)
        pending: list[dict] = []
        errors: list[dict[str, str]] = []
        seen_usernames: set[str] = set()

        for row_index, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            username = self._normalize_optional_text(cells[0])
            password = self._normalize_optional_text(cells[1])
            nickname = self._normalize_optional_text(cells[2])
            status_raw = self._normalize_optional_text(cells[3]) or UserStatusEnum.NORMAL.value
            roles_raw = cells[4]
            remark = self._normalize_optional_text(cells[5])

            if not username:
                errors.append({"row": row_index, "message": "用户名不能为空"})
                continue
            if username in seen_usernames:
                errors.append({"row": row_index, "message": "模版中存在重复的用户名"})
                continue
            seen_usernames.add(username)

            if not password:
                errors.append({"row": row_index, "message": "密码不能为空"})
                continue

            try:
                normalized_status = self._normalize_status(status_raw)
            except AppException as exc:
                errors.append({"row": row_index, "message": exc.msg})
                continue

            try:
                role_objects = self._resolve_roles_from_tokens(role_lookup, roles_raw)
            except AppException as exc:
                errors.append({"row": row_index, "message": exc.msg})
                continue

            pending.append(
                {
                    "row": row_index,
                    "username": username,
                    "password": password,
                    "nickname": nickname,
                    "status": normalized_status,
                    "roles": role_objects,
                    "remark": remark,
                }
            )

        if not pending:
            payload = {"created": 0, "failed": errors, "total": 0}
            return create_response("导入用户完成", payload, HTTP_STATUS_OK)

        existing_map = {
            user.username: user
            for user in user_crud.list_by_usernames(db, (item["username"] for item in pending))
        }

        created_count = 0
        for item in pending:
            if item["username"] in existing_map:
                errors.append({"row": item["row"], "message": "用户名已存在"})
                continue

            hashed_password = get_password_hash(item["password"])
            user_crud.create_with_roles(
                db,
                username=item["username"],
                hashed_password=hashed_password,
                nickname=item["nickname"],
                organization_id=None,
                status=item["status"],
                remark=item["remark"],
                roles=item["roles"],
            )
            created_count += 1

        payload = {"created": created_count, "failed": errors, "total": len(pending)}
        return create_response("导入用户完成", payload, HTTP_STATUS_OK)

    # ------------------------------------------------------------------
    # 内部辅助方法
    # ------------------------------------------------------------------

    def _serialize_user_summary(self, user: User) -> dict:
        return {
            "user_id": user.id,
            "username": user.username,
            "nickname": user.nickname,
            "status": user.status,
            "status_label": self._STATUS_LABELS.get(user.status, user.status),
            "role_ids": sorted(role.id for role in user.roles),
            "role_names": sorted(role.name for role in user.roles),
            "organization": self._serialize_organization(user.organization),
            "remark": user.remark,
            "create_time": format_datetime(user.create_time),
            "update_time": format_datetime(user.update_time),
            "is_active": user.is_active,
        }

    def _serialize_user_detail(self, user: User) -> dict:
        return self._serialize_user_summary(user)

    @staticmethod
    def _serialize_organization(organization: Optional[Organization]) -> Optional[dict]:
        if organization is None:
            return None
        return {"org_id": organization.id, "org_name": organization.name}

    def _normalize_statuses(self, statuses: Optional[Iterable[str]]) -> Optional[list[str]]:
        if statuses is None:
            return None
        normalized: list[str] = []
        for status in statuses:
            if not status:
                continue
            normalized.append(self._normalize_status(status))
        return normalized or None

    def _normalize_status(self, status: str) -> str:
        token = (status or "").strip().lower()
        if not token:
            raise AppException("用户状态不能为空", HTTP_STATUS_BAD_REQUEST)
        for code, label in self._STATUS_LABELS.items():
            if token == code:
                return code
            if token == label.lower():
                return code
        raise AppException("未知的用户状态", HTTP_STATUS_BAD_REQUEST)

    @staticmethod
    def _normalize_optional_text(value: Optional[str]) -> Optional[str]:
        if value is None:
            return None
        token = str(value).strip()
        return token or None

    def _load_roles(self, db: Session, role_ids: Optional[Iterable[int]]) -> Optional[list[Role]]:
        if role_ids is None:
            return None
        unique_order: list[int] = []
        seen: set[int] = set()
        for item in role_ids:
            if item is None:
                continue
            if item in seen:
                continue
            seen.add(item)
            unique_order.append(item)
        if not unique_order:
            return []

        roles = role_crud.list_by_ids(db, unique_order)
        missing = set(unique_order) - {role.id for role in roles}
        if missing:
            raise AppException(
                f"部分角色不存在：{', '.join(str(item) for item in sorted(missing))}",
                HTTP_STATUS_NOT_FOUND,
            )
        return roles

    def _load_organization(self, db: Session, organization_id: Optional[int]) -> Optional[Organization]:
        if organization_id is None:
            return None
        organization = organization_crud.get(db, organization_id)
        if organization is None:
            raise AppException("组织机构不存在", HTTP_STATUS_NOT_FOUND)
        return organization

    def _prepare_role_lookup(self, db: Session) -> dict[str, Role]:
        roles = role_crud.get_multi(db, skip=0, limit=10_000)
        lookup: dict[str, Role] = {}
        for role in roles:
            lookup[str(role.id)] = role
            lookup[role.name.lower()] = role
            if getattr(role, "role_key", None):
                lookup[role.role_key.lower()] = role
        return lookup

    def _resolve_roles_from_tokens(
        self,
        lookup: dict[str, Role],
        raw_tokens: Optional[object],
    ) -> list[Role]:
        if raw_tokens is None or raw_tokens == "":
            return []
        if isinstance(raw_tokens, (list, tuple)):
            tokens = raw_tokens
        else:
            tokens = str(raw_tokens).replace("，", ",").split(",")

        resolved: list[Role] = []
        seen: set[int] = set()
        for token in tokens:
            normalized = str(token).strip()
            if not normalized:
                continue
            candidate = lookup.get(normalized.lower())
            if candidate is None and normalized.isdigit():
                candidate = lookup.get(normalized)
            if candidate is None:
                raise AppException(f"角色无效：{normalized}", HTTP_STATUS_NOT_FOUND)
            if candidate.id in seen:
                continue
            seen.add(candidate.id)
            resolved.append(candidate)
        return resolved


user_service = UserService()
