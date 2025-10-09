"""角色管理接口的集成测试。"""

from __future__ import annotations

import io
import uuid

from fastapi.testclient import TestClient
from openpyxl import load_workbook


def _auth_headers(client: TestClient) -> dict[str, str]:
    login_resp = client.post(
        "/api/v1/auth/login",
        json={"username": "admin", "password": "admin123"},
    )
    token = login_resp.json()["data"]["access_token"]
    return {"Authorization": f"Bearer {token}"}


def _create_permission_node(client: TestClient, headers: dict[str, str]) -> int:
    unique_suffix = uuid.uuid4().hex[:8]

    root_resp = client.post(
        "/api/v1/access-controls",
        headers=headers,
        json={
            "name": f"角色测试菜单-{unique_suffix}",
            "type": "menu",
            "display_status": "show",
            "enabled_status": "enabled",
            "sort_order": 1,
        },
    )
    assert root_resp.status_code == 200
    root_id = root_resp.json()["data"]["id"]

    button_resp = client.post(
        "/api/v1/access-controls",
        headers=headers,
        json={
            "parent_id": root_id,
            "name": f"角色测试按钮-{unique_suffix}",
            "type": "button",
            "permission_code": f"system:role:test:{unique_suffix}",
            "enabled_status": "enabled",
            "sort_order": 1,
        },
    )
    assert button_resp.status_code == 200
    return button_resp.json()["data"]["id"]


def test_role_crud_flow_and_export(client: TestClient):
    """验证角色管理接口的增删改查与导出能力。"""
    headers = _auth_headers(client)
    permission_id = _create_permission_node(client, headers)

    rule_name = f"角色日志规则-{uuid.uuid4().hex[:6]}"
    monitor_resp = client.post(
        "/api/v1/logs/monitor-rules",
        headers=headers,
        json={
            "name": rule_name,
            "request_uri": "/api/v1/roles",
            "http_method": "ALL",
            "match_mode": "prefix",
            "is_enabled": True,
            "description": "测试角色接口日志采集",
            "operation_type_code": "query",
        },
    )
    assert monitor_resp.status_code == 201

    create_resp = client.post(
        "/api/v1/roles",
        headers=headers,
        json={
            "name": "测试角色A",
            "role_key": "role:test:a",
            "sort_order": 3,
            "status": "normal",
            "permission_ids": [permission_id],
            "remark": "用于测试",
        },
    )
    assert create_resp.status_code == 200
    create_payload = create_resp.json()
    role_id = create_payload["data"]["role_id"]
    assert create_payload["data"]["permission_ids"] == [permission_id]

    list_resp = client.get(
        "/api/v1/roles",
        headers=headers,
        params={"name": "测试角色A"},
    )
    assert list_resp.status_code == 200
    items = list_resp.json()["data"]["items"]
    assert any(item["role_id"] == role_id for item in items)

    status_resp = client.patch(
        f"/api/v1/roles/{role_id}/status",
        headers=headers,
        json={"status": "disabled"},
    )
    assert status_resp.status_code == 200
    assert status_resp.json()["data"]["status"] == "disabled"

    update_resp = client.put(
        f"/api/v1/roles/{role_id}",
        headers=headers,
        json={
            "name": "测试角色B",
            "role_key": "role:test:b",
            "sort_order": 5,
            "status": "normal",
            "permission_ids": [],
            "remark": "更新后",
        },
    )
    assert update_resp.status_code == 200
    assert update_resp.json()["data"]["role_name"] == "测试角色B"
    assert update_resp.json()["data"]["permission_ids"] == []

    detail_resp = client.get(f"/api/v1/roles/{role_id}", headers=headers)
    assert detail_resp.status_code == 200
    detail_payload = detail_resp.json()["data"]
    assert detail_payload["role_name"] == "测试角色B"
    assert detail_payload["status"] == "normal"

    export_resp = client.get(
        "/api/v1/roles/export",
        headers=headers,
        params={"statuses": ["normal"]},
    )
    assert export_resp.status_code == 200
    assert (
        export_resp.headers["Content-Type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(io.BytesIO(export_resp.content))
    sheet = workbook.active
    header = [cell.value for cell in sheet[1]]
    assert header == ["角色名称", "权限字符", "显示顺序", "状态", "创建时间"]
    exported_names = {sheet.cell(row=row, column=1).value for row in range(2, sheet.max_row + 1)}
    assert "测试角色B" in exported_names

    delete_resp = client.delete(f"/api/v1/roles/{role_id}", headers=headers)
    assert delete_resp.status_code == 200
    assert delete_resp.json()["data"]["role_id"] == role_id

    not_found_resp = client.get(f"/api/v1/roles/{role_id}", headers=headers)
    assert not_found_resp.status_code == 404
    assert not_found_resp.json()["msg"] == "角色不存在或已删除"

    log_resp = client.get(
        "/api/v1/logs/operations",
        headers=headers,
        params={"module": "角色管理"},
    )
    assert log_resp.status_code == 200
    log_items = log_resp.json()["data"]["items"]
    assert any(item["module"] == "角色管理" for item in log_items)


def test_role_name_and_key_unique(client: TestClient):
    """角色名称与权限字符应具备唯一性约束。"""
    headers = _auth_headers(client)
    permission_id = _create_permission_node(client, headers)

    base_payload = {
        "name": "唯一角色",
        "role_key": "role:unique",
        "sort_order": 1,
        "status": "normal",
        "permission_ids": [permission_id],
    }
    first_resp = client.post("/api/v1/roles", headers=headers, json=base_payload)
    assert first_resp.status_code == 200

    duplicate_name_resp = client.post(
        "/api/v1/roles",
        headers=headers,
        json=base_payload | {"role_key": "role:unique:duplicate"},
    )
    assert duplicate_name_resp.status_code == 409
    assert duplicate_name_resp.json()["msg"] == "角色名称已存在"

    duplicate_key_resp = client.post(
        "/api/v1/roles",
        headers=headers,
        json=base_payload | {"name": "唯一角色副本"},
    )
    assert duplicate_key_resp.status_code == 409
    assert duplicate_key_resp.json()["msg"] == "权限字符已存在"
